
# Bibliotecas importadas 

import os
import csv
from io import BytesIO
from datetime import datetime, timedelta

# renderização
from flask import (
    render_template, redirect, url_for, flash, send_file,
    request, session, Response, current_app as app
)
from werkzeug.utils import secure_filename


from app import db
from app.forms import ColaboradorForm, ProcessoForm, DeletarTodosForm, LoginForm, DeletarPorColaboradorForm
from app.models import Colaborador, Processo, DetalhesProcesso, SenhaUsuario, ImagemProcesso


from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader, simpleSplit

# openpyxl para gerar o objeto planilha
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.forms import TrocarSenhaForm
from app.models import SenhaUsuario

from flask import redirect

# ROTAS E DEFINIÇÕES

@app.route('/')
def home_redirect():
    return redirect(url_for('inicio'))

@app.route('/inicio')
def inicio():
    return render_template('inicio.html')

@app.route('/index')
def index():
    if not session.get('logado'):
        return redirect(url_for('login'))

    page = request.args.get('page', 1, type=int)
    processos = Processo.query.order_by(Processo.data_analise.desc()).paginate(page=page, per_page=10)

    return render_template('index.html', processos=processos, deletar_form=DeletarTodosForm())



@app.route('/colaborador/novo', methods=['GET', 'POST'])
def novo_colaborador():
    form = ColaboradorForm()
    if form.validate_on_submit():
        colaborador = Colaborador(nome=form.nome.data, email=form.email.data)
        db.session.add(colaborador)
        db.session.commit()
        flash('Colaborador adicionado com sucesso!', 'success')
        return redirect(url_for('index'))
    return render_template('novo_colaborador.html', form=form)


@app.route('/processo/novo', methods=['GET', 'POST'])
def novo_processo():
    form = ProcessoForm()
    form.colaborador.choices = [(c.id, c.nome) for c in Colaborador.query.all()]

    if form.validate_on_submit():

        imagens = request.files.getlist('imagem')

        brasil_time = datetime.utcnow() - timedelta(hours=3)

        filename_principal = None

        if imagens and imagens[0].filename != '':
            filename_principal = secure_filename(imagens[0].filename)
            caminho = os.path.join(app.config['UPLOAD_FOLDER'], filename_principal)
            imagens[0].save(caminho)

        processo = Processo(
            colaborador_id=form.colaborador.data,
            descricao=form.descricao.data,
            status=form.status.data,
            observacoes=form.observacoes.data,
            imagem=filename_principal,
            data_analise=brasil_time
        )

        db.session.add(processo)
        db.session.commit()

        # salvar todas as imagens
        for img in imagens:

            if img.filename == '':
                continue

            filename = secure_filename(img.filename)
            caminho = os.path.join(app.config['UPLOAD_FOLDER'], filename)

            if filename != filename_principal:
                img.save(caminho)

            nova_imagem = ImagemProcesso(
                processo_id=processo.id,
                nome_arquivo=filename
            )

            db.session.add(nova_imagem)

        # salvar detalhes
        detalhes = DetalhesProcesso(
            processo_id=processo.id,
            coordenacao=form.coordenacao.data,
            lideranca=form.lideranca.data,
            data_inspecao=form.data_inspecao.data,
            inspetor=form.inspetor.data,
            cpf_cnpj=form.cpf_cnpj.data,
            coop=form.coop.data,
            tipo=form.tipo.data,
            data_execucao=form.data_execucao.data,
            interno_externo=form.interno_externo.data,
            tipo_autorizacao=form.tipo_autorizacao.data,
            tipo_erro=form.tipo_erro.data,
            gravidade=form.gravidade.data,
            desvio_atencao=form.desvio_atencao.data,
            reversao=form.reversao.data,
            fluxos=form.fluxos.data,
            solicitante=form.solicitante.data
        )

        db.session.add(detalhes)

        # commit final
        db.session.commit()

        flash('Processo registrado com sucesso!', 'success')
        return redirect(url_for('index'))

    return render_template('novo_processo.html', form=form)

@app.route('/exportar-xlsx')
def exportar_xlsx():
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')

    if not data_inicio or not data_fim:
        flash('Informe o intervalo de datas para exportar.', 'warning')
        return redirect(url_for('index'))

    processos = Processo.query.filter(
        Processo.data_analise.between(data_inicio, data_fim)
    ).order_by(Processo.data_analise.asc()).all()

    wb = Workbook()
    ws = wb.active

    titulo = f"{data_inicio}_a_{data_fim}"
    ws.title = titulo[:31]

    # Estilos (verifica a possibilidade de alterar a cor do cabeçalho para azul pp a planilha da qualidade é azul, (pesquisar tom de azul))
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

#cores para cabeçalho
# ADD8E6 ou 0000FF ou 87CEEB ou 4682B4  


    # Cabeçalhos da planilha em ordem (de acordo com a planilha da qualidade, deve ser nessa ordem)
    cabecalhos = [
        "COORDENACAO", "LIDERANCA", "DATA DA INSPECAO", "INSPETOR", "CPF/CNPJ", "COOP", "TIPO",
        "DATA DE EXECUCAO", "INTERNO OU EXTERNO", "TIPO DE AUTORIZACAO", "COLABORADOR",
        "TIPO DE ERRO", "DESCRICAO", "GRAVIDADE", "DESVIO DE ATENCAO", "REVERSAO",
        "QUAIS FLUXOS", "OBSERVACOES", "SOLICITANTE"
    ]
    ws.append(cabecalhos)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border

    for p in processos:
        d = p.detalhes
        linha = [
            d.coordenacao if d else '',
            d.lideranca if d else '',
            d.data_inspecao.strftime('%d/%m/%Y') if d and d.data_inspecao else '',
            d.inspetor if d else '',
            d.cpf_cnpj if d else '',
            d.coop if d else '',
            d.tipo if d else '',
            d.data_execucao.strftime('%d/%m/%Y') if d and d.data_execucao else '',
            d.interno_externo if d else '',
            d.tipo_autorizacao if d else '',
            p.colaborador.nome,
            d.tipo_erro if d else '',
            p.descricao,
            d.gravidade if d else '',
            d.desvio_atencao if d else '',
            d.reversao if d else '',
            d.fluxos if d else '',
            p.observacoes or '',
            d.solicitante if d else ''
        ]
        ws.append([str(c) if c is not None else '' for c in linha])  

    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(cabecalhos)):
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    # mesma largura utiliza na planilha da qualidade! (Ajustar caso necessário)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    
    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'processos_{data_inicio}_a_{data_fim}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# DELETAR DATOS   (revisar se possivel melhorar rota)

@app.route('/deletar-registros', methods=['GET', 'POST'])
def deletar_registros():

    form = DeletarPorColaboradorForm()

    # carregar colaboradores no select
    form.colaborador.choices = [(c.id, c.nome) for c in Colaborador.query.all()]

    if form.validate_on_submit():

        colaborador_id = form.colaborador.data

        processos = Processo.query.filter_by(colaborador_id=colaborador_id).all()

        for processo in processos:

            # deletar detalhes primeiro
            if processo.detalhes:
                db.session.delete(processo.detalhes)

            # deletar processo
            db.session.delete(processo)

        db.session.commit()

        flash('Processos do colaborador foram removidos com sucesso.', 'warning')

        return redirect(url_for('index'))

    return render_template('deletar_registros.html', form=form)
# PAINEL 

@app.route('/painel', methods=['GET'])
def painel():
    colaboradores = Colaborador.query.all()

    colaborador_id = request.args.get('colaborador', type=int)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    status = request.args.get('status')
    palavra = request.args.get('palavra')
    page = request.args.get('page', 1, type=int)

    query = Processo.query

    if colaborador_id:
        query = query.filter_by(colaborador_id=colaborador_id)
    
    if status:
        query = query.filter_by(status=status)
    
    if palavra:
        query = query.filter(Processo.descricao.ilike(f"%{palavra}%"))

    if data_inicio:
        query = query.filter(Processo.data_analise >= data_inicio)
    if data_fim:
        query = query.filter(Processo.data_analise <= data_fim)

    total = query.count()
    conformes = query.filter_by(status='Conforme').count()
    desvios = query.filter_by(status='Desvio').count()

    processos = query.order_by(Processo.data_analise.desc()).paginate(page=page, per_page=10)

    return render_template('painel.html',
                           colaboradores=colaboradores,
                           colaborador_id=colaborador_id,
                           total=total,
                           conformes=conformes,
                           desvios=desvios,
                           processos=processos)



# LOGIN  (critico, precisa de revisão e se possivel melhorar)

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        email = form.email.data
        senha = form.senha.data

        senha_registrada = SenhaUsuario.query.filter_by(email=email).first()

        if senha_registrada:
            # Se já existe senha trocada, só aceita com hash
            if senha_registrada.verificar_senha(senha):
                session['logado'] = True
                flash('Login realizado com sucesso!', 'success')
                return redirect(url_for('index'))
            else:
                flash('Senha incorreta.', 'danger')
        else:
            # Só aceita senha padrão se ainda não foi trocada
            colaborador = Colaborador.query.filter_by(email=email).first()
            if colaborador and senha == '123456':
                session['logado'] = True
                flash('Login com senha padrão.', 'warning')
                return redirect(url_for('index'))
            else:
                flash('Senha incorreta ou colaborador inexistente.', 'danger')

    return render_template('login.html', form=form)


@app.route('/logout')
def logout():
    session.pop('logado', None)
    flash('Sessão encerrada com sucesso.', 'info')
    return redirect(url_for('login'))


@app.route('/processo/<int:id>/pdf')
def gerar_pdf(id):

    processo = Processo.query.get_or_404(id)

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)

    y = 800
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(50, y, "PARECER DE INSPEÇÃO DE PROCESSO")
    y -= 30

    pdf.setFont("Helvetica", 12)

    # Usa data_execucao se existir, senão usa data_analise
    data_execucao = (
        processo.detalhes.data_execucao.strftime('%d/%m/%Y')
        if processo.detalhes and processo.detalhes.data_execucao
        else processo.data_analise.strftime('%d/%m/%Y %H:%M')
    )

    pdf.drawString(50, y, f"Data: {data_execucao}")
    y -= 20

    pdf.drawString(50, y, f"Colaborador: {processo.colaborador.nome}")
    y -= 20

    pdf.drawString(50, y, f"Status: {processo.status}")
    y -= 20

    pdf.drawString(50, y, f"Descrição: {processo.descricao}")
    y -= 40


    # Observações
    if processo.observacoes:

        pdf.drawString(50, y, "Observações:")
        y -= 20

        linhas_obs = simpleSplit(processo.observacoes, "Helvetica", 12, 500)

        text = pdf.beginText(50, y)
        text.setFont("Helvetica", 12)

        for linha in linhas_obs:
            text.textLine(linha)

        pdf.drawText(text)
        y = text.getY() - 20


    # Parecer quando é desvio
    if processo.status == 'Desvio':

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y, "PARECER:")
        y -= 20

        pdf.setFont("Helvetica", 12)

        parecer = (
            "Conforme análise realizada, identificamos um ponto de atenção neste processo. "
            "Recomenda-se avaliação complementar e plano de ação conforme aplicável."
        )

        linhas_parecer = simpleSplit(parecer, "Helvetica", 12, 500)

        text = pdf.beginText(50, y)

        for linha in linhas_parecer:
            text.textLine(linha)

        pdf.drawText(text)
        y = text.getY() - 20


    # IMAGENS
    if processo.imagens:

        for img in processo.imagens:

            imagem_path = os.path.join(
                app.root_path,
                'static',
                'uploads',
                img.nome_arquivo
            )

            if os.path.exists(imagem_path):

                try:

                    if y < 250:
                        pdf.showPage()
                        y = 800

                    pdf.drawImage(
                        ImageReader(imagem_path),
                        130,
                        y - 200,
                        width=300,
                        height=200
                    )

                    y -= 220

                except Exception:
                    pass

    else:
        pdf.drawString(50, y, "Nenhuma evidência anexada.")


    pdf.showPage()
    pdf.save()

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'processo_{id}.pdf',
        mimetype='application/pdf'
    )
@app.route('/trocar-senha', methods=['GET', 'POST'])
def trocar_senha():
    form = TrocarSenhaForm()
    if form.validate_on_submit():
        email = form.email.data
        senha_atual = form.senha_atual.data
        nova_senha = form.nova_senha.data

        senha_registrada = SenhaUsuario.query.filter_by(email=email).first()

        if senha_registrada:
            # Se existe hash, validar senha atual
            if not senha_registrada.verificar_senha(senha_atual):
                flash('Senha atual incorreta.', 'danger')
                return render_template('trocar_senha.html', form=form)
        else:
            # Se ainda não existe hash, só aceita padrão '123456'
            if senha_atual != '123456':
                flash('Senha atual incorreta.', 'danger')
                return render_template('trocar_senha.html', form=form)

            # Cria registro novo na tabela
            senha_registrada = SenhaUsuario(email=email)
            db.session.add(senha_registrada)

        # Salva nova senha
        senha_registrada.set_senha(nova_senha)
        db.session.commit()
        flash('Senha alterada com sucesso!', 'success')
        return redirect(url_for('login'))

    return render_template('trocar_senha.html', form=form)

# MEU PERFIL (utiliza o filter, sinceramente poderia ser melhor)

@app.route('/meu_perfil')
def meu_perfil():
    hoje = datetime.today()
    inicio_mes_atual = hoje.replace(day=1)
    inicio_mes_anterior = (inicio_mes_atual - timedelta(days=1)).replace(day=1)
    fim_mes_anterior = inicio_mes_atual - timedelta(days=1)

    processos_atual = Processo.query.filter(Processo.data_analise >= inicio_mes_atual).all()
    processos_anterior = Processo.query.filter(Processo.data_analise.between(inicio_mes_anterior, fim_mes_anterior)).all()

    total_mes_atual = len(processos_atual)
    conformes_mes_atual = len([p for p in processos_atual if p.status == 'Conforme'])
    desvios_mes_atual = len([p for p in processos_atual if p.status == 'Desvio'])

    total_mes_anterior = len(processos_anterior)
    conformes_mes_anterior = len([p for p in processos_anterior if p.status == 'Conforme'])
    desvios_mes_anterior = len([p for p in processos_anterior if p.status == 'Desvio'])

    diff_total = total_mes_atual - total_mes_anterior
    diff_conformes = conformes_mes_atual - conformes_mes_anterior
    diff_desvios = desvios_mes_atual - desvios_mes_anterior

    return render_template('meu_perfil.html',
                           total_mes_atual=total_mes_atual,
                           conformes_mes_atual=conformes_mes_atual,
                           desvios_mes_atual=desvios_mes_atual,
                           diff_total=diff_total,
                           diff_conformes=diff_conformes,
                           diff_desvios=diff_desvios)




@app.route('/colaboradores')
def listar_colaboradores():
    colaboradores = Colaborador.query.all()

    dados_colaboradores = []
    for colaborador in colaboradores:
        total_processos = Processo.query.filter_by(colaborador_id=colaborador.id).count()
        dados_colaboradores.append({
            'id': colaborador.id,
            'nome': colaborador.nome,
            'email': colaborador.email,
            'total_processos': total_processos
        })

    return render_template('colaboradores.html', colaboradores=dados_colaboradores)


@app.route('/grafico-processos')
def grafico_processos():
    conformes = Processo.query.filter_by(status='Conforme').count()
    desvios = Processo.query.filter_by(status='Desvio').count()

    return render_template('grafico_processos.html',
                           conformes=conformes,
                           desvios=desvios)


#redirecionamento suporte com assistente

@app.route('/assistente-virtual')
def assistente_virtual():
    return redirect("https://nicolebr1.onrender.com/")


#nova rota
@app.route('/processo/<int:id>/imagens')
def ver_imagens_processo(id):

    processo = Processo.query.get_or_404(id)

    return render_template(
        'ver_imagens.html',
        processo=processo
    )